"""
Excel export utility for generating reports.
Uses openpyxl to create professional Excel spreadsheets.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from decimal import Decimal


class ExcelExporter:
    """
    Base class for Excel exports with common styling.
    """
    
    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        
        # Define color scheme (Modern Industrial)
        self.colors = {
            'primary': '059669',      # Emerald 600
            'background': 'F8FAFC',   # Slate 50
            'text': '0F172A',         # Slate 900
            'border': 'E2E8F0',       # Slate 200
            'success': '10B981',      # Green 500
            'warning': 'F59E0B',      # Amber 500
            'error': 'EF4444',        # Red 500
        }
        
        # Define styles
        self.header_font = Font(name='Inter', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color=self.colors['primary'], 
                                       end_color=self.colors['primary'], 
                                       fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center')
        
        self.title_font = Font(name='Inter', size=14, bold=True, color=self.colors['primary'])
        self.normal_font = Font(name='Inter', size=10, color=self.colors['text'])
        self.bold_font = Font(name='Inter', size=10, bold=True, color=self.colors['text'])
        
        self.border = Border(
            left=Side(style='thin', color=self.colors['border']),
            right=Side(style='thin', color=self.colors['border']),
            top=Side(style='thin', color=self.colors['border']),
            bottom=Side(style='thin', color=self.colors['border'])
        )
    
    def style_header_row(self, row_num, columns):
        """
        Apply header styling to a row.
        """
        for col_num in range(1, columns + 1):
            cell = self.ws.cell(row=row_num, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
    
    def auto_adjust_columns(self):
        """
        Auto-adjust column widths based on content.
        """
        for column in self.ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            self.ws.column_dimensions[column_letter].width = adjusted_width
    
    def save(self, filename):
        """
        Save workbook to file.
        """
        self.auto_adjust_columns()
        self.wb.save(filename)
        return filename


class InvoiceExporter(ExcelExporter):
    """
    Export invoices to Excel.
    """
    
    def generate(self, invoices, filename):
        """
        Generate Excel file with invoice data.
        """
        self.ws.title = "Invoices"
        
        # Title
        self.ws['A1'] = 'INVOICE REPORT'
        self.ws['A1'].font = self.title_font
        self.ws['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
        self.ws['A2'].font = Font(name='Inter', size=9, color=self.colors['text'])
        
        # Headers
        headers = [
            'Invoice #',
            'Unit',
            'Tenant',
            'Billing Period',
            'Invoice Date',
            'Due Date',
            'Water (m³)',
            'Water Charge (KES)',
            'Electricity (kWh)',
            'Electricity Charge (KES)',
            'Fixed Charges (KES)',
            'Subtotal (KES)',
            'Previous Balance (KES)',
            'Total Due (KES)',
            'Status'
        ]
        
        header_row = 4
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=header_row, column=col_num)
            cell.value = header
        
        self.style_header_row(header_row, len(headers))
        
        # Data rows
        row_num = header_row + 1
        for invoice in invoices:
            data = [
                invoice.invoice_number,
                invoice.unit.unit_number,
                invoice.tenant.user.get_full_name() or invoice.tenant.user.username,
                invoice.billing_period,
                invoice.invoice_date.strftime('%Y-%m-%d'),
                invoice.due_date.strftime('%Y-%m-%d'),
                float(invoice.water_units) if invoice.water_units else 0,
                float(invoice.water_charge) if invoice.water_charge else 0,
                float(invoice.electricity_units) if invoice.electricity_units else 0,
                float(invoice.electricity_charge) if invoice.electricity_charge else 0,
                float(invoice.total_fixed_charges) if invoice.total_fixed_charges else 0,
                float(invoice.subtotal) if invoice.subtotal else 0,
                float(invoice.previous_balance) if invoice.previous_balance else 0,
                float(invoice.total_due) if invoice.total_due else 0,
                invoice.get_status_display()
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = self.ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                
                # Number formatting for currency columns
                if col_num in [8, 10, 11, 12, 13, 14]:
                    cell.number_format = '#,##0.00'
                
                # Color-code status
                if col_num == 15:
                    if value == 'Paid':
                        cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
                    elif value == 'Overdue':
                        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                    elif value == 'Partially Paid':
                        cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            
            row_num += 1
        
        # Summary row
        summary_row = row_num + 1
        self.ws[f'A{summary_row}'] = 'TOTALS'
        self.ws[f'A{summary_row}'].font = self.bold_font
        
        # Calculate totals
        total_water_charge = sum(float(inv.water_charge or 0) for inv in invoices)
        total_elec_charge = sum(float(inv.electricity_charge or 0) for inv in invoices)
        total_fixed = sum(float(inv.total_fixed_charges or 0) for inv in invoices)
        total_subtotal = sum(float(inv.subtotal or 0) for inv in invoices)
        total_due = sum(float(inv.total_due or 0) for inv in invoices)
        
        self.ws[f'H{summary_row}'] = total_water_charge
        self.ws[f'J{summary_row}'] = total_elec_charge
        self.ws[f'K{summary_row}'] = total_fixed
        self.ws[f'L{summary_row}'] = total_subtotal
        self.ws[f'N{summary_row}'] = total_due
        
        # Format totals
        for col in ['H', 'J', 'K', 'L', 'N']:
            cell = self.ws[f'{col}{summary_row}']
            cell.font = self.bold_font
            cell.number_format = '#,##0.00'
            cell.fill = PatternFill(start_color=self.colors['background'], 
                                   end_color=self.colors['background'], 
                                   fill_type='solid')
        
        return self.save(filename)


class PaymentExporter(ExcelExporter):
    """
    Export payments to Excel.
    """
    
    def generate(self, payments, filename):
        """
        Generate Excel file with payment data.
        """
        self.ws.title = "Payments"
        
        # Title
        self.ws['A1'] = 'PAYMENT REPORT'
        self.ws['A1'].font = self.title_font
        self.ws['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
        self.ws['A2'].font = Font(name='Inter', size=9, color=self.colors['text'])
        
        # Headers
        headers = [
            'Payment Date',
            'Invoice #',
            'Unit',
            'Tenant',
            'Amount Paid (KES)',
            'Payment Method',
            'M-Pesa Reference',
            'Phone Number',
            'Recorded By',
            'Notes'
        ]
        
        header_row = 4
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=header_row, column=col_num)
            cell.value = header
        
        self.style_header_row(header_row, len(headers))
        
        # Data rows
        row_num = header_row + 1
        for payment in payments:
            data = [
                payment.payment_date.strftime('%Y-%m-%d'),
                payment.invoice.invoice_number,
                payment.invoice.unit.unit_number,
                payment.invoice.tenant.user.get_full_name() or payment.invoice.tenant.user.username,
                float(payment.amount_paid),
                payment.get_payment_method_display(),
                payment.mpesa_reference or '',
                payment.mpesa_phone or '',
                payment.recorded_by.username if payment.recorded_by else 'System',
                payment.notes or ''
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = self.ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                
                # Number formatting for currency
                if col_num == 5:
                    cell.number_format = '#,##0.00'
            
            row_num += 1
        
        # Summary row
        summary_row = row_num + 1
        self.ws[f'A{summary_row}'] = 'TOTAL COLLECTED'
        self.ws[f'A{summary_row}'].font = self.bold_font
        
        total_collected = sum(float(payment.amount_paid) for payment in payments)
        self.ws[f'E{summary_row}'] = total_collected
        self.ws[f'E{summary_row}'].font = self.bold_font
        self.ws[f'E{summary_row}'].number_format = '#,##0.00'
        self.ws[f'E{summary_row}'].fill = PatternFill(start_color=self.colors['background'], 
                                                       end_color=self.colors['background'], 
                                                       fill_type='solid')
        
        return self.save(filename)


class ConsumptionExporter(ExcelExporter):
    """
    Export consumption data to Excel.
    """
    
    def generate(self, readings, filename):
        """
        Generate Excel file with consumption data.
        """
        self.ws.title = "Consumption"
        
        # Title
        self.ws['A1'] = 'CONSUMPTION REPORT'
        self.ws['A1'].font = self.title_font
        self.ws['A2'] = f'Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
        self.ws['A2'].font = Font(name='Inter', size=9, color=self.colors['text'])
        
        # Headers
        headers = [
            'Date',
            'Unit',
            'Estate',
            'Meter Type',
            'Meter Number',
            'Reading Value',
            'Previous Reading',
            'Consumption',
            'Anomaly',
            'Anomaly Type',
            'Recorded By'
        ]
        
        header_row = 4
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=header_row, column=col_num)
            cell.value = header
        
        self.style_header_row(header_row, len(headers))
        
        # Data rows
        row_num = header_row + 1
        for reading in readings:
            data = [
                reading.reading_date.strftime('%Y-%m-%d'),
                reading.meter.unit.unit_number,
                reading.meter.unit.estate_name,
                reading.meter.get_meter_type_display(),
                reading.meter.meter_number,
                float(reading.reading_value),
                float(reading.previous_reading) if reading.previous_reading else 0,
                float(reading.consumption),
                'Yes' if reading.has_anomaly else 'No',
                reading.anomaly_type or '',
                reading.recorded_by.username if reading.recorded_by else ''
            ]
            
            for col_num, value in enumerate(data, 1):
                cell = self.ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.font = self.normal_font
                cell.border = self.border
                
                # Number formatting
                if col_num in [6, 7, 8]:
                    cell.number_format = '#,##0.00'
                
                # Highlight anomalies
                if col_num == 9 and value == 'Yes':
                    cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                    cell.font = Font(name='Inter', size=10, bold=True, color=self.colors['warning'])
            
            row_num += 1
        
        # Summary
        summary_row = row_num + 1
        self.ws[f'A{summary_row}'] = 'TOTAL CONSUMPTION'
        self.ws[f'A{summary_row}'].font = self.bold_font
        
        total_consumption = sum(float(reading.consumption) for reading in readings)
        self.ws[f'H{summary_row}'] = total_consumption
        self.ws[f'H{summary_row}'].font = self.bold_font
        self.ws[f'H{summary_row}'].number_format = '#,##0.00'
        self.ws[f'H{summary_row}'].fill = PatternFill(start_color=self.colors['background'], 
                                                       end_color=self.colors['background'], 
                                                       fill_type='solid')
        
        return self.save(filename)