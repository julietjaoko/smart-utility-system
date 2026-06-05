from decimal import Decimal
from datetime import timedelta
from io import BytesIO
from unittest.mock import Mock, patch

from django.urls import reverse
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook

from .models import (
    AccountBalance,
    AuditLog,
    Invoice,
    Meter,
    MeterReading,
    Payment,
    PropertyManager,
    Tenant,
    TenantPreferences,
    Unit,
    User,
)
from .sms_utils import AfricasTalkingSMS
from .views import recalculate_tenant_ledger


class BillingLedgerTests(TestCase):
    def setUp(self):
        manager_user = User.objects.create_user(
            username='manager@example.com',
            password='password',
            role='PROPERTY_MANAGER',
        )
        self.manager = PropertyManager.objects.create(
            user=manager_user,
            estate_name='Test Estate',
        )
        self.unit = Unit.objects.create(
            unit_number='A1',
            estate_name='Test Estate',
            manager=self.manager,
        )
        tenant_user = User.objects.create_user(
            username='tenant@example.com',
            password='password',
            role='TENANT',
        )
        self.tenant = Tenant.objects.create(user=tenant_user, unit=self.unit)

    def make_invoice(self, number, subtotal, previous_balance, days_offset=0):
        today = timezone.now().date()
        return Invoice.objects.create(
            unit=self.unit,
            tenant=self.tenant,
            invoice_number=number,
            invoice_date=today + timedelta(days=days_offset),
            due_date=today + timedelta(days=days_offset + 10),
            billing_period=number,
            subtotal=Decimal(subtotal),
            previous_balance=Decimal(previous_balance),
            total_due=Decimal(subtotal) + Decimal(previous_balance),
            generated_by=self.manager.user,
        )

    def test_late_payment_reduces_newer_invoice_carried_balance(self):
        old_invoice = self.make_invoice('INV-OLD', '1000.00', '0.00')
        new_invoice = self.make_invoice('INV-NEW', '500.00', '1000.00')
        AccountBalance.objects.create(tenant=self.tenant, current_balance=Decimal('1500.00'))

        Payment.objects.create(
            invoice=old_invoice,
            payment_date=timezone.now().date(),
            amount_paid=Decimal('1000.00'),
            payment_method='CASH',
        )

        current_balance = recalculate_tenant_ledger(self.tenant)
        old_invoice.refresh_from_db()
        new_invoice.refresh_from_db()

        self.assertEqual(old_invoice.status, 'PAID')
        self.assertEqual(new_invoice.previous_balance, Decimal('0.00'))
        self.assertEqual(new_invoice.total_due, Decimal('500.00'))
        self.assertEqual(current_balance, Decimal('500.00'))
        self.assertEqual(new_invoice.balance_due, Decimal('500.00'))

    def test_latest_invoice_payment_clears_carried_previous_invoices(self):
        carried_balance = Decimal('0.00')
        old_invoices = []

        for index in range(6):
            invoice = self.make_invoice(
                f'INV-OLD-{index + 1}',
                '3000.00',
                str(carried_balance),
                days_offset=index,
            )
            old_invoices.append(invoice)
            carried_balance += Decimal('3000.00')

        latest_invoice = self.make_invoice(
            'INV-LATEST',
            '7000.00',
            str(carried_balance),
            days_offset=6,
        )
        AccountBalance.objects.create(tenant=self.tenant, current_balance=Decimal('25000.00'))

        Payment.objects.create(
            invoice=latest_invoice,
            payment_date=timezone.now().date(),
            amount_paid=Decimal('25000.00'),
            payment_method='CASH',
        )

        current_balance = recalculate_tenant_ledger(self.tenant)

        for invoice in old_invoices:
            invoice.refresh_from_db()
            self.assertEqual(invoice.status, 'PAID')

        latest_invoice.refresh_from_db()
        self.assertEqual(latest_invoice.status, 'PAID')
        self.assertEqual(latest_invoice.previous_balance, Decimal('18000.00'))
        self.assertEqual(latest_invoice.total_due, Decimal('25000.00'))
        self.assertEqual(current_balance, Decimal('0.00'))

    def test_manager_dashboard_loads_with_insights(self):
        self.client.force_login(self.manager.user)

        response = self.client.get(reverse('manager_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Insights')


class MpesaStkStatusTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(
            username='manager-mpesa@example.com',
            password='password',
            role='PROPERTY_MANAGER',
        )
        self.manager = PropertyManager.objects.create(
            user=self.manager_user,
            estate_name='Mpesa Estate',
        )
        self.unit = Unit.objects.create(
            unit_number='M1',
            estate_name='Mpesa Estate',
            manager=self.manager,
        )
        self.tenant_user = User.objects.create_user(
            username='tenant-mpesa@example.com',
            password='password',
            role='TENANT',
        )
        self.tenant = Tenant.objects.create(
            user=self.tenant_user,
            unit=self.unit,
            phone_number='0712345678',
        )
        self.invoice = Invoice.objects.create(
            unit=self.unit,
            tenant=self.tenant,
            invoice_number='INV-MPESA-001',
            invoice_date=timezone.now().date(),
            due_date=timezone.now().date() + timedelta(days=7),
            billing_period='June 2026',
            subtotal=Decimal('1000.00'),
            total_due=Decimal('1000.00'),
            generated_by=self.manager_user,
        )

    @patch('core.mpesa.MpesaDarajaSandbox.initiate_stk_push')
    def test_stk_initiation_returns_status_url(self, mock_stk_push):
        mock_stk_push.return_value = {
            'success': True,
            'checkout_request_id': 'ws_CO_test',
            'merchant_request_id': 'mr_test',
            'response_description': 'Success. Request accepted for processing',
        }
        self.client.force_login(self.tenant_user)

        response = self.client.post(
            reverse('initiate_mpesa_payment', args=[self.invoice.id]),
            {'phone_number': '0712345678'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['success'])
        self.assertEqual(payload['checkout_request_id'], 'ws_CO_test')
        self.assertEqual(
            payload['status_url'],
            reverse('mpesa_payment_status', args=[self.invoice.id]),
        )

    @patch('core.mpesa.MpesaDarajaSandbox.query_stk_status')
    def test_stk_status_returns_failed_prompt_quickly(self, mock_status):
        mock_status.return_value = {
            'success': False,
            'status': 'FAILED',
            'result_code': 1032,
            'result_desc': 'Request cancelled by user',
        }
        self.client.force_login(self.tenant_user)

        response = self.client.get(
            reverse('mpesa_payment_status', args=[self.invoice.id]),
            {'checkout_request_id': 'ws_CO_test'},
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertFalse(payload['success'])
        self.assertEqual(payload['status'], 'FAILED')
        self.assertEqual(payload['result_code'], 1032)
        self.assertEqual(payload['message'], 'Request cancelled by user')

    def test_tenant_mpesa_ajax_form_does_not_trigger_global_page_loader(self):
        self.client.force_login(self.tenant_user)

        response = self.client.get(reverse('tenant_invoices'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<form id="mpesa-payment-form" data-no-loader>')
        self.assertContains(response, 'window.SUMSLoader')


class ConsumptionExportTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(
            username='manager-export@example.com',
            password='password',
            role='PROPERTY_MANAGER',
        )
        self.manager = PropertyManager.objects.create(
            user=self.manager_user,
            estate_name='Export Estate',
        )
        self.unit = Unit.objects.create(
            unit_number='B1',
            estate_name='Export Estate',
            manager=self.manager,
        )
        self.meter = Meter.objects.get(unit=self.unit, meter_type='WATER')

    def test_meter_readings_export_generates_workbook_with_previous_reading(self):
        first_date = timezone.now() - timedelta(days=30)
        second_date = timezone.now()
        MeterReading.objects.create(
            meter=self.meter,
            reading_value=Decimal('100.00'),
            reading_date=first_date,
            recorded_by=self.manager_user,
        )
        MeterReading.objects.create(
            meter=self.meter,
            reading_value=Decimal('145.00'),
            reading_date=second_date,
            recorded_by=self.manager_user,
        )

        self.client.force_login(self.manager_user)
        response = self.client.get(reverse('export_consumption_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )

        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook['Consumption']
        self.assertEqual(worksheet['B5'].value, 'B1')
        self.assertEqual(worksheet['F5'].value, 145)
        self.assertEqual(worksheet['G5'].value, 100)
        self.assertEqual(worksheet['H5'].value, 45)


class ConsumptionAnalyticsFinalReadingTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(
            username='manager-analytics-final@example.com',
            password='password',
            role='PROPERTY_MANAGER',
        )
        self.manager = PropertyManager.objects.create(
            user=self.manager_user,
            estate_name='Analytics Estate',
        )
        self.unit = Unit.objects.create(
            unit_number='A2',
            estate_name='Analytics Estate',
            manager=self.manager,
        )
        self.meter = Meter.objects.get(unit=self.unit, meter_type='WATER')
        now = timezone.now()

        verified = MeterReading.objects.create(
            meter=self.meter,
            reading_value=Decimal('20.00'),
            reading_date=now - timedelta(days=10),
            recorded_by=self.manager_user,
        )
        MeterReading.objects.filter(pk=verified.pk).update(
            consumption=Decimal('20.00'),
            is_anomaly=False,
            anomaly_type='',
            verification_status='VERIFIED',
        )

        pending = MeterReading.objects.create(
            meter=self.meter,
            reading_value=Decimal('1020.00'),
            reading_date=now - timedelta(days=5),
            recorded_by=self.manager_user,
        )
        MeterReading.objects.filter(pk=pending.pk).update(
            consumption=Decimal('1000.00'),
            is_anomaly=True,
            anomaly_type='hard_limit_exceeded',
            verification_status='PENDING',
        )

    def test_consumption_analytics_totals_use_only_verified_readings(self):
        self.client.force_login(self.manager_user)

        response = self.client.get(reverse('consumption_analytics'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['total_consumption'], 20.0)
        self.assertEqual(response.context['chart_data'], '[20.0]')
        self.assertEqual(response.context['anomaly_count'], 1)

    def test_reports_center_consumption_totals_use_only_verified_readings(self):
        from .reporting.services import build_consumption_report

        today = timezone.now().date()

        report = build_consumption_report(
            self.manager,
            today - timedelta(days=30),
            today + timedelta(days=1),
        )

        self.assertEqual(report['total_consumption'], Decimal('20.00'))
        self.assertEqual(report['reading_count'], 1)
        self.assertEqual(report['anomaly_count'], 1)

    def test_advanced_analytics_excludes_negative_consumption_from_final_metrics(self):
        negative = MeterReading.objects.create(
            meter=self.meter,
            reading_value=Decimal('10.00'),
            reading_date=timezone.now() - timedelta(days=2),
            recorded_by=self.manager_user,
        )
        MeterReading.objects.filter(pk=negative.pk).update(
            consumption=Decimal('-10.00'),
            is_anomaly=True,
            anomaly_type='negative_consumption',
            verification_status='VERIFIED',
        )

        self.client.force_login(self.manager_user)
        response = self.client.get(reverse('advanced_analytics'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['water_consumption'], Decimal('20.00'))
        self.assertNotIn('-10.0', response.context['water_monthly'])


class TenantConsumptionAlertTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(
            username='manager-alerts@example.com',
            password='password',
            role='PROPERTY_MANAGER',
        )
        self.manager = PropertyManager.objects.create(
            user=self.manager_user,
            estate_name='Alert Estate',
        )
        self.unit = Unit.objects.create(
            unit_number='C1',
            estate_name='Alert Estate',
            manager=self.manager,
        )
        self.tenant_user = User.objects.create_user(
            username='tenant-alerts@example.com',
            password='password',
            role='TENANT',
        )
        self.tenant = Tenant.objects.create(user=self.tenant_user, unit=self.unit)
        self.meter = Meter.objects.get(unit=self.unit, meter_type='WATER')

    def create_high_usage_reading(self):
        now = timezone.now()
        values = [
            ('10.00', now - timedelta(days=40)),
            ('20.00', now - timedelta(days=30)),
            ('30.00', now - timedelta(days=20)),
            ('50.00', now - timedelta(days=10)),
        ]
        for value, reading_date in values:
            MeterReading.objects.create(
                meter=self.meter,
                reading_value=Decimal(value),
                reading_date=reading_date,
                recorded_by=self.manager_user,
            )

    def test_tenant_dashboard_shows_high_consumption_alert_when_enabled(self):
        TenantPreferences.objects.create(
            tenant=self.tenant,
            show_consumption_alerts=True,
        )
        self.create_high_usage_reading()

        self.client.force_login(self.tenant_user)
        response = self.client.get(reverse('tenant_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'High Consumption Alert')
        self.assertContains(response, 'High')

    def test_tenant_dashboard_hides_high_consumption_alert_when_disabled(self):
        TenantPreferences.objects.create(
            tenant=self.tenant,
            show_consumption_alerts=False,
        )
        self.create_high_usage_reading()

        self.client.force_login(self.tenant_user)
        response = self.client.get(reverse('tenant_dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'High Consumption Alert')


class LoginMessageTests(TestCase):
    def test_invalid_credentials_show_only_generic_login_error(self):
        response = self.client.post(
            reverse('login'),
            {'username': 'unknown', 'password': 'wrong'},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Invalid username or password.')
        self.assertNotContains(response, 'profile not found')

    def test_missing_profile_does_not_show_internal_login_error(self):
        User.objects.create_user(
            username='manager-without-profile',
            password='password123',
            role='PROPERTY_MANAGER',
        )

        response = self.client.post(
            reverse('login'),
            {'username': 'manager-without-profile', 'password': 'password123'},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Invalid username or password.')
        self.assertNotContains(response, 'Property Manager profile not found')

    def test_successful_login_does_not_show_login_success_message(self):
        manager_user = User.objects.create_user(
            username='manager-login-clean@example.com',
            password='password123',
            role='PROPERTY_MANAGER',
        )
        PropertyManager.objects.create(
            user=manager_user,
            estate_name='Clean Login Estate',
        )

        response = self.client.post(
            reverse('login'),
            {'username': 'manager-login-clean@example.com', 'password': 'password123'},
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'Logged in successfully.')

    def test_authenticated_pages_render_django_messages_as_toasts(self):
        manager_user = User.objects.create_user(
            username='manager-toast-owner@example.com',
            password='password123',
            role='PROPERTY_MANAGER',
        )
        manager = PropertyManager.objects.create(
            user=manager_user,
            estate_name='Toast Estate',
        )
        unit = Unit.objects.create(
            unit_number='T1',
            estate_name='Toast Estate',
            manager=manager,
        )
        tenant_user = User.objects.create_user(
            username='tenant-toast@example.com',
            password='password123',
            role='TENANT',
        )
        Tenant.objects.create(
            user=tenant_user,
            unit=unit,
            phone_number='0712345678',
        )

        self.client.force_login(tenant_user)
        response = self.client.post(
            reverse('tenant_preferences'),
            {
                'phone_number': '0712345678',
                'enable_sms_notifications': 'on',
                'enable_email_notifications': 'on',
                'show_consumption_alerts': 'on',
            },
            follow=True,
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="toast success"')
        self.assertContains(response, 'Preferences updated successfully.')


class AfricasTalkingSMSTests(TestCase):
    def test_success_depends_on_recipient_status(self):
        service = AfricasTalkingSMS()
        service.sms = Mock()
        service.sms.send.return_value = {
            'SMSMessageData': {
                'Message': 'Sent to 1/1 Total Cost: KES 0.8000',
                'Recipients': [
                    {
                        'statusCode': 101,
                        'number': '+254712345678',
                        'status': 'Success',
                        'messageId': 'ATXid_test',
                    }
                ],
            }
        }

        result = service.send_sms('0712345678', 'Test message')

        self.assertTrue(result['success'])
        self.assertEqual(result['status'], 'Success')
        service.sms.send.assert_called_once_with(
            message='Test message',
            recipients=['+254712345678'],
        )

    def test_failed_recipient_status_is_reported_as_failed(self):
        service = AfricasTalkingSMS()
        service.sms = Mock()
        service.sms.send.return_value = {
            'SMSMessageData': {
                'Message': 'Sent to 0/1 Total Cost: KES 0.0000',
                'Recipients': [
                    {
                        'statusCode': 406,
                        'number': '+254712345678',
                        'status': 'InvalidPhoneNumber',
                    }
                ],
            }
        }

        result = service.send_sms('0712345678', 'Test message')

        self.assertFalse(result['success'])
        self.assertEqual(result['status'], 'InvalidPhoneNumber')
        self.assertEqual(result['error'], 'InvalidPhoneNumber')


class ReportsCenterExportTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(
            username='manager-report-export@example.com',
            password='password',
            role='PROPERTY_MANAGER',
        )
        self.manager = PropertyManager.objects.create(
            user=self.manager_user,
            estate_name='Report Estate',
        )
        self.unit = Unit.objects.create(
            unit_number='R1',
            estate_name='Report Estate',
            manager=self.manager,
        )
        self.tenant_user = User.objects.create_user(
            username='tenant-report-export@example.com',
            password='password',
            role='TENANT',
        )
        self.tenant = Tenant.objects.create(user=self.tenant_user, unit=self.unit)
        self.meter = Meter.objects.get(unit=self.unit, meter_type='WATER')

        self.invoice = Invoice.objects.create(
            unit=self.unit,
            tenant=self.tenant,
            invoice_number='INV-REPORT-001',
            invoice_date=timezone.now().date(),
            due_date=timezone.now().date() - timedelta(days=5),
            billing_period='June 2026',
            subtotal=Decimal('1000.00'),
            total_due=Decimal('1000.00'),
            status='OVERDUE',
            generated_by=self.manager_user,
        )
        Payment.objects.create(
            invoice=self.invoice,
            payment_date=timezone.now().date(),
            amount_paid=Decimal('250.00'),
            payment_method='CASH',
            recorded_by=self.manager_user,
        )
        MeterReading.objects.create(
            meter=self.meter,
            reading_value=Decimal('200.00'),
            reading_date=timezone.now(),
            consumption=Decimal('80.00'),
            is_anomaly=True,
            anomaly_type='HIGH_USAGE',
            verification_status='PENDING',
            recorded_by=self.manager_user,
        )

    def test_all_report_center_excel_exports_download_workbooks(self):
        self.client.force_login(self.manager_user)

        for report_type in ('financial', 'arrears', 'consumption', 'anomalies'):
            with self.subTest(report_type=report_type):
                response = self.client.get(
                    reverse('export_report_excel'),
                    {'report': report_type},
                )

                self.assertEqual(response.status_code, 200)
                self.assertEqual(
                    response['Content-Type'],
                    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                workbook = load_workbook(BytesIO(response.content))
                self.assertGreaterEqual(len(workbook.sheetnames), 1)


class ActivityLogExportTests(TestCase):
    def setUp(self):
        self.manager_user = User.objects.create_user(
            username='manager-log-export@example.com',
            password='password',
            role='PROPERTY_MANAGER',
        )
        self.manager = PropertyManager.objects.create(
            user=self.manager_user,
            estate_name='Log Estate',
        )
        self.admin_user = User.objects.create_user(
            username='admin-log-export@example.com',
            password='password',
            role='SYSTEM_ADMIN',
        )
        AuditLog.objects.create(
            actor=self.manager_user,
            property_manager=self.manager,
            category='AUTH',
            action='LOGIN',
            message='Manager signed in',
        )

    def test_manager_activity_log_export_downloads_workbook(self):
        self.client.force_login(self.manager_user)

        response = self.client.get(reverse('export_activity_logs_excel'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook['Activity Log']
        self.assertEqual(worksheet['A1'].value, 'SYSTEM ACTIVITY LOG')
        self.assertEqual(worksheet['C5'].value, 'Authentication')
        self.assertEqual(worksheet['D5'].value, 'LOGIN')

    def test_system_admin_activity_log_export_downloads_workbook(self):
        self.client.force_login(self.admin_user)

        response = self.client.get(reverse('system_admin_export_activity_logs'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content))
        worksheet = workbook['Activity Log']
        self.assertEqual(worksheet['A1'].value, 'SYSTEM ACTIVITY LOG')
        self.assertEqual(worksheet['C4'].value, 'Property Manager')
        self.assertEqual(worksheet['D5'].value, 'Authentication')
        self.assertEqual(worksheet['E5'].value, 'LOGIN')
